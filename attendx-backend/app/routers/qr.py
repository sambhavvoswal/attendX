from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from typing import Dict, Any, List
import io

from app.dependencies import get_current_user
from app.services import firebase_service
from app.services.sheets_service import SheetsService

router = APIRouter(prefix="/api/qr", tags=["qr"])
sheets_service = SheetsService()

import csv
import openpyxl

@router.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    try:
        contents = await file.read()
        
        headers = []
        rows = []
        
        if file.filename.endswith('.csv'):
            decoded_file = contents.decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            headers = reader.fieldnames if reader.fieldnames else []
            for row in reader:
                rows.append({k: (v if v is not None else "") for k, v in row.items()})
        else:
            # Excel fallback using openpyxl
            wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
            sheet = wb.active
            
            # Extract headers from the first row
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            if header_row:
                headers = [str(h) for h in header_row[0] if h is not None]
                
                # Extract the rest of the rows
                for row_tuple in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    has_data = False
                    for idx, header in enumerate(headers):
                        val = ""
                        if idx < len(row_tuple) and row_tuple[idx] is not None:
                            val = str(row_tuple[idx])
                            has_data = True
                        row_dict[header] = val
                    if has_data:
                        rows.append(row_dict)
            
        return {
            "headers": headers,
            "rows": rows
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

